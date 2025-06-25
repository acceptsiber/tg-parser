import json
import re
from typing import Any, Dict, List, Optional, Union
import asyncio
import os
from telethon import TelegramClient
from telethon.tl.types import MessageMediaPhoto
import logging
import requests
import time

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from json_parser import parse_message_to_json

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# Загружаем JSON шаблон
template_json = {
    "type": "1",
    "user": "608",
    "object_type": "5",
    "object_fields": [
        {
            "id": "896",
            "title": "Цена",
            "field_type": "number",
            "data": {"value": "5000000"},
            "options": [],
            "required": True,
        },
        {
            "id": "951",
            "title": "Общая площадь",
            "field_type": "number",
            "data": {"value": "Узбекистан, Ташкент, Циолковский"},
            "options": [],
            "required": False,
        },
        {
            "id": "869",
            "title": "Адрес объекта",
            "field_type": "address",
            "data": {"value": "Узбекистан, Ташкент, Циолковский"},
            "options": [],
            "required": False,
        },
        {
            "id": "871",
            "title": "Описание",
            "field_type": "rich_text",
            "data": {"value": "1231"},
            "options": [],
            "required": False,
        },
        {
            "id": "883",
            "title": "Тип сделки",
            "field_type": "object_type",
            "data": {"value": "1"},
            "options": [],
            "required": True,
        },
    ],
}

API_URL = "https://hata.uz/api/object/"
BEARER_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ0b2tlbl90eXBlIjoiYWNjZXNzIiwiZXhwIjoxNzUwODMxMTQwLCJpYXQiOjE3NTA3NDQ3NDAsImp0aSI6IjRiNWExYjE0ZjExNjQwZGJiYjM3ZDQ1MGRjODJjMjAyIiwidXNlcl9pZCI6NjA4fQ.5gxo_jCwsxiR2MukLlQfE6p7plz7T0SU9OA9HOMNuFM"

# Заголовки для запроса
headers = {
    "Authorization": f"Bearer {BEARER_TOKEN}",
    # "Content-Type": "multipart/form-data",
}


class TelegramParser:
    def __init__(self, client: TelegramClient):
        self.client = client

    async def get_messages(
        self, entity: str | int, limit: int = 100
    ) -> List[Dict[str, str]]:
        """Gets messages with media files."""
        messages: List[Dict[str, str]] = []
        messages_data = await self.client.get_messages(entity, limit=limit)
        for message in messages_data:

            first_line = message.text.split("\n")[0].strip()
            message_id = (
                first_line.split()[0]
                if first_line and len(first_line.split()) > 0
                else "No ID"
            )

            messages.append(
                {
                    "telegram_id": message.id,
                    "id": re.sub(r"\*", "", message_id),
                    "message": re.sub(r"\s+", " ", message.message).strip(),
                    "photo": message.media,
                }
            )

        return messages

    async def group_objects(
        self, objects
    ) -> List[Dict[str, Union[str, List[MessageMediaPhoto]]]]:
        grouped = []
        current_group = {"message": "", "photos": []}

        for obj in objects:
            try:
                if obj["message"]:
                    current_group["message"] = obj
                    current_group["photos"].append(obj["photo"])
                    grouped.append(current_group)
                    current_group = {"message": "", "photos": []}
                else:
                    current_group["photos"].append(obj["photo"])
            except:
                continue

        return grouped

    async def download_photos(
        self, grouped_message: List[Dict[str, Union[str, List[MessageMediaPhoto]]]]
    ):
        os.makedirs("photos", exist_ok=True)
        tasks = []
        for message in grouped_message:
            for photo in message.get("photos"):
                if not photo:
                    continue

                # Проверяем тип медиа
                if isinstance(photo, MessageMediaPhoto):
                    try:
                        filename = os.path.join(
                            "photos",
                            f"{photo.photo.id}.jpg",
                        )
                        task = asyncio.create_task(
                            self.client.download_media(photo, file=filename)
                        )
                        tasks.append(task)
                    except AttributeError as e:
                        logging.warning(f"Skipping photo due to error: {e}")
                        continue

        return await asyncio.gather(*tasks)

    async def set_photo_id_to_message(
        self, grouped_message: List[Dict[str, Union[str, List[MessageMediaPhoto]]]]
    ):
        for message in grouped_message:
            message["photos_id"] = [photo.photo.id for photo in message["photos"]]

        return grouped_message


async def create_excel_with_ids_and_text(
    data: List[Dict[str, Union[str, List[MessageMediaPhoto]]]],
    filename: str = "tg_messages.xlsx",
    photos_dir: str = "photos",
) -> None:

    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Текст", "Фото"])

    # Собираем все фото и создаем гиперссылки
    if os.path.exists(photos_dir):
        for item in data:
            message_id = item["message"]["id"]
            row = [message_id, item["message"]["message"], ""]  # Поле для гиперссылок
            ws.append(row)

            # Ищем все фото для этого ID
            photo_counter = 1
            for photo_file in os.listdir(photos_dir):
                if photo_file.startswith(f"{message_id}_"):
                    photo_path = os.path.abspath(os.path.join(photos_dir, photo_file))

                    # Создаем кликабельную ссылку
                    link_text = f"Фото {photo_counter}"
                    cell = ws.cell(row=ws.max_row, column=3)
                    cell.value = link_text
                    cell.hyperlink = Hyperlink(
                        ref=cell.coordinate,
                        target=photo_path,
                        tooltip=f"Открыть {photo_file}",
                    )
                    photo_counter += 1

    # Настраиваем ширину колонок
    ws.column_dimensions[get_column_letter(1)].width = 15  # ID
    ws.column_dimensions[get_column_letter(2)].width = 50  # Текст
    ws.column_dimensions[get_column_letter(3)].width = 25  # Фото

    wb.save(filename)
    print(f"✅ Файл {filename} создан с кликабельными ссылками!")


def send_objects(json_list):
    for index, json_obj in enumerate(json_list, 1):
        try:
            print(f"Отправка объекта {index} из {len(json_list)}...")

            response = requests.post(
                API_URL,
                json={
                    "type": "1",
                    "user": "608",
                    "object_type": "2",
                    "object_fields": [
                        {
                            "id": "896",
                            "title": "Цена",
                            "field_type": "number",
                            "data": {"value": "5000000"},
                            "options": [],
                            "required": True,
                        },
                        {
                            "id": "883",
                            "title": "Тип сделки",
                            "field_type": "object_type",
                            "data": {"value": "1"},
                            "options": [],
                            "required": True,
                        },
                    ],
                },
                headers=headers,
                auth=BEARER_TOKEN,
            )

            if response.status_code == 200 or response.status_code == 201:
                print(
                    f"Успешно отправлен объект {index}. Ответ сервера: {response.text}"
                )
            else:
                print(
                    f"Ошибка при отправке объекта {index}. Код статуса: {response.status_code}"
                )
                print(f"Ответ сервера: {response}")

            # Небольшая задержка между запросами, чтобы не перегружать сервер
            time.sleep(0.1)

        except Exception as e:
            print(f"Ошибка при отправке объекта {index}: {str(e)}")

        print("-" * 50)


def send_json_as_multipart(json_data, field_name="fields", headers={}):
    """
    Отправляет JSON данные как файл в multipart/form-data запросе.

    Args:
        api_url: URL API-сервиса.
        json_data: JSON данные для отправки (словарь или список словарей).
        field_name: Имя поля в multipart/form-data, которое будет содержать JSON. Defaults to "fields".
        headers: Дополнительные заголовки запроса.

    Returns:
        Объект response от requests.
    """
    # Преобразуем JSON данные в строку
    json_string = json.dumps(
        json_data, ensure_ascii=False
    )  # ensure_ascii=False чтобы не было проблем с кириллицей

    # Создаем файл в памяти (не нужно записывать на диск)
    files = {
        field_name: (
            "data.json",
            json_string,
            "application/json",
        )
    }

    try:
        response = requests.post(API_URL, files=files, headers=headers)
        response.raise_for_status()
        print(response.raise_for_status())
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при отправке запроса: {e}")
        return None


async def main():
    api_id = 23128708
    api_hash = "ee3dfa7067eb520a05bfc749083f9ab0"
    phone_number = "+79130028603"
    chat_entity = "prodajadomov_mirzo_ulugbek"

    async with TelegramClient("s", api_id, api_hash) as client:
        try:
            await client.connect()
            if not await client.is_user_authorized():
                await client.send_code_request(phone_number)
                await client.sign_in(phone_number, input("Enter code: "))

            parser = TelegramParser(client)

            messages = await parser.get_messages(chat_entity, limit=32)
            parsed_messages = []

            groups = await parser.group_objects(messages)

            for message in groups:
                parsed_message = parse_message_to_json(
                    message["message"]["message"], template_json
                )

                send_json_as_multipart(json_data=parsed_message, headers=headers)
                parsed_messages.append(parsed_message)

            # await parser.download_photos(groups)
            # groups = await parser.set_photo_id_to_message(groups)

            # await create_excel_with_ids_and_text(groups)
            # print(*groups, sep="\n\n")
            # print(*parsed_messages, sep="\n\n")
            # print(*groups, sep="\n\n")

        except Exception as e:
            logging.error(f"An error occurred: {e}")
        finally:
            await client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
